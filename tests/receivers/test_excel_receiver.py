import asyncio
import inspect
from datetime import datetime, timedelta
from pathlib import Path
from typing import AsyncGenerator

import dask.dataframe as dd
from dask.dataframe.utils import assert_eq
import pandas as pd
import pytest
from openpyxl import load_workbook

from etl_core.receivers.files.excel.excel_receiver import ExcelReceiver
from etl_core.metrics.component_metrics.component_metrics import ComponentMetrics


@pytest.fixture
def metrics() -> ComponentMetrics:
    return ComponentMetrics(
        started_at=datetime.now(),
        processing_time=timedelta(0),
        error_count=0,
        lines_received=0,
        lines_forwarded=0,
    )


@pytest.fixture
def sample_excel_file() -> Path:
    return (
        Path(__file__).parent.parent
        / "components"
        / "files"
        / "data"
        / "test_data.xlsx"
    )


@pytest.mark.asyncio
async def test_excelreceiver_read_row_streaming_incremental(
    sample_excel_file: Path, metrics: ComponentMetrics
) -> None:
    r = ExcelReceiver()
    rows = r.read_row(filepath=sample_excel_file, metrics=metrics)

    assert inspect.isasyncgen(rows) or isinstance(rows, AsyncGenerator)

    rec1 = await asyncio.wait_for(anext(rows), timeout=0.5)
    assert rec1["name"] == "Alice"
    assert metrics.lines_received == 1
    assert metrics.error_count == 0

    rec2 = await asyncio.wait_for(anext(rows), timeout=0.5)
    assert rec2["name"] == "Bob"
    assert metrics.lines_received == 2
    assert metrics.error_count == 0

    rec3 = await asyncio.wait_for(anext(rows), timeout=0.5)
    assert rec3["name"] == "Charlie"
    assert metrics.lines_received == 3
    assert metrics.error_count == 0

    with pytest.raises(StopAsyncIteration):
        await asyncio.wait_for(anext(rows), timeout=0.2)

    await rows.aclose()


@pytest.mark.asyncio
async def test_read_excel_bulk(
    sample_excel_file: Path, metrics: ComponentMetrics
) -> None:
    r = ExcelReceiver()
    df = await r.read_bulk(filepath=sample_excel_file, metrics=metrics)
    assert isinstance(df, pd.DataFrame)
    assert len(df) == 3
    assert set(df.columns) == {"id", "name"}
    assert set(df["name"]) == {"Alice", "Bob", "Charlie"}
    assert metrics.lines_received == 3
    assert metrics.error_count == 0


@pytest.mark.asyncio
async def test_read_excel_bigdata_matches_expected(
    sample_excel_file: Path, metrics: ComponentMetrics
) -> None:
    r = ExcelReceiver()
    ddf = await r.read_bigdata(filepath=sample_excel_file, metrics=metrics)
    assert isinstance(ddf, dd.DataFrame)

    expected = pd.read_excel(sample_excel_file)

    assert_eq(
        ddf.set_index(None).compute().sort_values(by=["name"]).reset_index(drop=True),
        expected.sort_values(by=["name"]).reset_index(drop=True),
        check_dtype=False,
        check_index=False,
    )

    assert metrics.lines_received == len(expected)
    assert metrics.error_count == 0


@pytest.mark.asyncio
async def test_write_excel_row(tmp_path: Path, metrics: ComponentMetrics) -> None:
    file_path = tmp_path / "out_row.xlsx"
    r = ExcelReceiver()

    await r.write_row(
        filepath=file_path, metrics=metrics, row={"id": "10", "name": "Daisy"}
    )
    await r.write_row(
        filepath=file_path, metrics=metrics, row={"id": "11", "name": "Eli"}
    )

    assert file_path.exists()

    wb = load_workbook(filename=file_path)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert set(headers) == {"id", "name"}

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))

    assert len(data) == 2
    names = {str(rec["name"]) for rec in data}
    assert names == {"Daisy", "Eli"}

    ids = {str(rec["id"]) for rec in data}
    assert ids == {"10", "11"}

    assert metrics.error_count == 0
    assert metrics.lines_forwarded == 2


@pytest.mark.asyncio
async def test_write_excel_bulk(tmp_path: Path, metrics: ComponentMetrics) -> None:
    file_path = tmp_path / "out_bulk.xlsx"
    r = ExcelReceiver()

    data = pd.DataFrame(
        [
            {"id": "20", "name": "Finn"},
            {"id": "21", "name": "Gina"},
        ]
    )

    await r.write_bulk(filepath=file_path, metrics=metrics, data=data)

    assert file_path.exists()

    wb = load_workbook(filename=file_path)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert set(headers) == {"id", "name"}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))

    assert len(rows) == 2
    names = {str(rw["name"]) for rw in rows}
    ids = {str(rw["id"]) for rw in rows}
    assert names == {"Finn", "Gina"}
    assert ids == {"20", "21"}

    assert metrics.error_count == 0
    assert metrics.lines_forwarded == 2


@pytest.mark.asyncio
async def test_write_excel_bigdata(tmp_path: Path, metrics: ComponentMetrics) -> None:
    file_path = tmp_path / "out_big.xlsx"
    r = ExcelReceiver()

    pdf = pd.DataFrame(
        [
            {"id": 30, "name": "Hugo"},
            {"id": 31, "name": "Ivy"},
        ]
    )
    ddf_in = dd.from_pandas(pdf, npartitions=1)

    await r.write_bigdata(filepath=file_path, metrics=metrics, data=ddf_in)

    assert file_path.exists()

    wb = load_workbook(filename=file_path)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert set(headers) == {"id", "name"}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))

    assert len(rows) == 2
    names = {str(rw["name"]) for rw in rows}
    ids = {str(rw["id"]) for rw in rows}
    assert names == {"Hugo", "Ivy"}
    assert ids == {"30", "31"}

    assert metrics.error_count == 0
    assert metrics.lines_forwarded == 2


@pytest.mark.asyncio
async def test_excelreceiver_missing_file_bulk_raises(
    metrics: ComponentMetrics, tmp_path: Path
) -> None:
    r = ExcelReceiver()
    with pytest.raises(FileNotFoundError):
        await r.read_bulk(filepath=tmp_path / "missing.xlsx", metrics=metrics)


@pytest.mark.asyncio
async def test_excelreceiver_invalid_file_raises(
    metrics: ComponentMetrics, tmp_path: Path
) -> None:
    invalid = tmp_path / "invalid.xlsx"
    invalid.write_text("not an excel file")
    r = ExcelReceiver()
    with pytest.raises(Exception):
        await r.read_bulk(filepath=invalid, metrics=metrics)
