import asyncio
import inspect
from datetime import datetime, timedelta
from pathlib import Path
from typing import AsyncGenerator, Any, Dict, List

import dask.dataframe as dd
import pandas as pd
import pytest
from openpyxl import load_workbook

from etl_core.components.file_components.excel.read_excel import ReadExcel
from etl_core.components.file_components.excel.write_excel import WriteExcel
from etl_core.metrics.component_metrics.component_metrics import ComponentMetrics
from etl_core.strategies.row_strategy import RowExecutionStrategy
from etl_core.strategies.bulk_strategy import BulkExecutionStrategy
from etl_core.strategies.bigdata_strategy import BigDataExecutionStrategy


DATA_DIR = Path(__file__).parent / "data"
VALID_XLSX = DATA_DIR / "test_data.xlsx"
MISSING_VALUES_XLSX = DATA_DIR / "test_data_missing.xlsx"
SCHEMA_MISMATCH_XLSX = DATA_DIR / "test_data_schema.xlsx"
WRONG_TYPES_XLSX = DATA_DIR / "test_data_types.xlsx"


@pytest.fixture
def metrics() -> ComponentMetrics:
    return ComponentMetrics(
        started_at=datetime.now(),
        processing_time=timedelta(0),
        error_count=0,
        lines_received=0,
        lines_forwarded=0,
    )


@pytest.mark.asyncio
async def test_readexcel_valid_bulk(metrics: ComponentMetrics) -> None:
    comp = ReadExcel(
        name="ReadExcel_Bulk_Valid",
        description="Valid Excel file",
        comp_type="read_excel",
        filepath=VALID_XLSX,
    )
    comp.strategy = BulkExecutionStrategy()

    gen = comp.execute(payload=None, metrics=metrics)
    assert inspect.isasyncgen(gen) or isinstance(gen, AsyncGenerator)

    dfs: List[pd.DataFrame] = []
    async for df in gen:
        dfs.append(df)

    assert len(dfs) == 1
    df = dfs[0]
    assert isinstance(df, pd.DataFrame)
    assert len(df) == 3
    assert set(df.columns) == {"id", "name"}
    assert set(df["name"]) == {"Alice", "Bob", "Charlie"}
    assert metrics.error_count == 0
    assert metrics.lines_received == 3


@pytest.mark.asyncio
async def test_readexcel_missing_values_bulk(metrics: ComponentMetrics) -> None:
    comp = ReadExcel(
        name="ReadExcel_Bulk_Missing",
        description="Missing values",
        comp_type="read_excel",
        filepath=MISSING_VALUES_XLSX,
    )
    comp.strategy = BulkExecutionStrategy()

    gen = comp.execute(payload=None, metrics=metrics)
    async for df in gen:
        assert df.isna().any().any()

    assert metrics.error_count == 0


@pytest.mark.asyncio
async def test_readexcel_wrong_types_bulk(metrics: ComponentMetrics) -> None:
    comp = ReadExcel(
        name="ReadExcel_Bulk_WrongTypes",
        description="Wrong types",
        comp_type="read_excel",
        filepath=WRONG_TYPES_XLSX,
    )
    comp.strategy = BulkExecutionStrategy()

    gen = comp.execute(payload=None, metrics=metrics)
    async for df in gen:
        assert {"id", "name"}.issubset(df.columns)
        # Ensure coercion to string is possible for id
        df["id"].astype(str)

    assert metrics.error_count == 0


@pytest.mark.asyncio
async def test_readexcel_row_streaming(metrics: ComponentMetrics) -> None:
    comp = ReadExcel(
        name="ReadExcel_Row_Stream",
        description="Row streaming",
        comp_type="read_excel",
        filepath=VALID_XLSX,
    )
    comp.strategy = RowExecutionStrategy()

    rows = comp.execute(payload=None, metrics=metrics)
    assert inspect.isasyncgen(rows) or isinstance(rows, AsyncGenerator)

    collected: List[Dict[str, Any]] = []
    for _ in range(3):
        rec = await asyncio.wait_for(anext(rows), timeout=0.5)
        collected.append(rec)

    await rows.aclose()

    assert collected[0]["name"] == "Alice"
    assert collected[1]["name"] == "Bob"
    assert collected[2]["name"] == "Charlie"
    assert metrics.lines_received == 3
    assert metrics.error_count == 0


@pytest.mark.asyncio
async def test_readexcel_bigdata(metrics: ComponentMetrics) -> None:
    comp = ReadExcel(
        name="ReadExcel_BigData",
        description="Read Excel with Dask",
        comp_type="read_excel",
        filepath=VALID_XLSX,
    )
    comp.strategy = BigDataExecutionStrategy()

    gen = comp.execute(payload=None, metrics=metrics)
    ddfs: List[dd.DataFrame] = []
    async for ddf in gen:
        ddfs.append(ddf)

    assert len(ddfs) == 1
    pdf = ddfs[0].compute()
    assert set(pdf["name"]) == {"Alice", "Bob", "Charlie"}
    assert metrics.error_count == 0
    assert metrics.lines_received == 3


@pytest.mark.asyncio
async def test_writeexcel_row(tmp_path: Path, metrics: ComponentMetrics) -> None:
    out_fp = tmp_path / "single.xlsx"
    comp = WriteExcel(
        name="WriteExcel_Row",
        description="Write single row",
        comp_type="write_excel",
        filepath=out_fp,
    )
    comp.strategy = RowExecutionStrategy()

    row1 = {"id": "1", "name": "Zoe"}
    row2 = {"id": "2", "name": "Liam"}
    await anext(comp.execute(payload=row1, metrics=metrics), None)
    await anext(comp.execute(payload=row2, metrics=metrics), None)

    assert out_fp.exists()

    wb = load_workbook(filename=out_fp)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert set(headers) == {"id", "name"}

    data: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, r)))

    assert len(data) == 2
    names = {str(rec["name"]) for rec in data}
    ids = {str(rec["id"]) for rec in data}
    assert names == {"Zoe", "Liam"}
    assert ids == {"1", "2"}

    assert metrics.error_count == 0
    assert metrics.lines_forwarded == 2


@pytest.mark.asyncio
async def test_writeexcel_bulk(tmp_path: Path, metrics: ComponentMetrics) -> None:
    out_fp = tmp_path / "bulk.xlsx"
    comp = WriteExcel(
        name="WriteExcel_Bulk",
        description="Write multiple rows",
        comp_type="write_excel",
        filepath=out_fp,
    )
    comp.strategy = BulkExecutionStrategy()

    data = pd.DataFrame(
        [
            {"id": "1", "name": "A"},
            {"id": "2", "name": "B"},
            {"id": "3", "name": "C"},
        ]
    )

    await anext(comp.execute(payload=data, metrics=metrics), None)
    assert out_fp.exists()

    wb = load_workbook(filename=out_fp)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert set(headers) == {"id", "name"}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, r)))

    assert len(rows) == 3
    assert {str(x["name"]) for x in rows} == {"A", "B", "C"}

    assert metrics.error_count == 0
    assert metrics.lines_forwarded == 3


@pytest.mark.asyncio
async def test_writeexcel_bigdata(tmp_path: Path, metrics: ComponentMetrics) -> None:
    out_fp = tmp_path / "big.xlsx"
    comp = WriteExcel(
        name="WriteExcel_BigData",
        description="Write Dask DataFrame",
        comp_type="write_excel",
        filepath=out_fp,
    )
    comp.strategy = BigDataExecutionStrategy()

    ddf_in = dd.from_pandas(
        pd.DataFrame(
            [
                {"id": "10", "name": "Nina"},
                {"id": "11", "name": "Omar"},
            ]
        ),
        npartitions=1,
    )

    await anext(comp.execute(payload=ddf_in, metrics=metrics), None)
    assert out_fp.exists()

    wb = load_workbook(filename=out_fp)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert set(headers) == {"id", "name"}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, r)))

    assert len(rows) == 2
    assert {str(x["name"]) for x in rows} == {"Nina", "Omar"}
    assert {str(x["id"]) for x in rows} == {"10", "11"}

    assert metrics.error_count == 0
    assert metrics.lines_forwarded == 2


@pytest.mark.asyncio
async def test_readexcel_invalid_file(
    metrics: ComponentMetrics, tmp_path: Path
) -> None:
    invalid = tmp_path / "invalid.xlsx"
    invalid.write_text("not a real excel file")
    comp = ReadExcel(
        name="ReadExcel_Invalid",
        description="Invalid Excel file",
        comp_type="read_excel",
        filepath=invalid,
    )
    comp.strategy = BulkExecutionStrategy()

    gen = comp.execute(payload=None, metrics=metrics)
    with pytest.raises(Exception):
        await anext(gen)


@pytest.mark.asyncio
async def test_readexcel_missing_file(
    metrics: ComponentMetrics, tmp_path: Path
) -> None:
    missing = tmp_path / "missing.xlsx"
    comp = ReadExcel(
        name="ReadExcel_Missing",
        description="Missing file",
        comp_type="read_excel",
        filepath=missing,
    )
    comp.strategy = BulkExecutionStrategy()

    gen = comp.execute(payload=None, metrics=metrics)
    with pytest.raises(FileNotFoundError):
        await anext(gen)
