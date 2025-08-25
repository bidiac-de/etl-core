from pathlib import Path
from typing import Dict, Iterator, List, Optional, Sequence, Union, Tuple

import dask.dataframe as dd
import pandas as pd

from src.etl_core.receivers.files.file_helper import (
    ensure_file_exists,
    resolve_file_path,
)

# File extension constants
XLSX_EXT = ".xlsx"
XLSM_EXT = ".xlsm"
XLS_EXT = ".xls"

READABLE_EXTS = {XLSX_EXT, XLSM_EXT, XLS_EXT}
WRITABLE_EXTS = {XLSX_EXT, XLSM_EXT}

SHEET_DEFAULT = "Sheet1"


def _ext(path: Path) -> str:
    return path.suffix.lower()


def _engine_for_read(ext: str) -> str:
    if ext in {XLSX_EXT, XLSM_EXT}:
        return "openpyxl"
    if ext == XLS_EXT:
        return "xlrd"
    raise ValueError(f"Unsupported excel extension for read: '{ext}'.")


def _engine_for_write(ext: str) -> str:
    if ext in WRITABLE_EXTS:
        return "openpyxl"
    raise ValueError(
        f"Writing '{ext}' is not supported. Use one of: {sorted(WRITABLE_EXTS)}."
    )


def _prepare_read(path: Path) -> Tuple[Path, str]:
    path = resolve_file_path(path)
    ensure_file_exists(path)
    return path, _engine_for_read(_ext(path))


def _prepare_write(path: Path) -> Tuple[Path, str]:
    path = resolve_file_path(Path(path))
    path.parent.mkdir(parents=True, exist_ok=True)
    return path, _engine_for_write(_ext(path))


def _df_replace_nans_inplace(df: pd.DataFrame) -> pd.DataFrame:
    return df.where(pd.notna(df), None)


def _read_openpyxl_header(rows_iter) -> List[str]:
    """Consume rows_iter until a non-empty header row is found; return header."""
    for first in rows_iter:
        if first and any(v is not None and v != "" for v in first):
            return [str(h) if h is not None else "" for h in first]
    raise ValueError("No header row found in worksheet.")


def _iter_openpyxl_rows(
    path: Path, sheet_name: Optional[str]
) -> Iterator[Dict[str, object]]:
    """Yield row dicts from an openpyxl worksheet without loading entire sheet."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        headers = _read_openpyxl_header(rows)
        for vals in rows:
            if vals is None:
                continue
            vals_list = list(vals[: len(headers)])
            if len(vals_list) < len(headers):
                vals_list += [None] * (len(headers) - len(vals_list))
            yield dict(zip(headers, vals_list))
    finally:
        wb.close()


def _iter_xlrd_rows(
    path: Path, sheet_name: Optional[str]
) -> Iterator[Dict[str, object]]:
    """Yield row dicts from an xlrd sheet (.xls)."""
    import xlrd  # xlrd>=2 only supports .xls

    book = xlrd.open_workbook(path)
    sh = book.sheet_by_name(sheet_name) if sheet_name else book.sheet_by_index(0)
    if sh.nrows == 0:
        raise ValueError("Worksheet is empty.")
    headers = list(
        str(sh.cell_value(0, c)) if sh.cell_value(0, c) != "" else ""
        for c in range(sh.ncols)
    )
    if not any(h for h in headers):
        raise ValueError("No header row found in worksheet.")
    for r in range(1, sh.nrows):
        vals = [sh.cell_value(r, c) for c in range(sh.ncols)]
        yield dict(zip(headers, vals))


def read_excel_rows(
    path: Path, sheet_name: Optional[str] = None
) -> Iterator[Dict[str, object]]:
    """
    Stream rows as dicts without loading entire sheet into memory.
    First non-empty row is treated as header.
    """
    path, engine = _prepare_read(path)
    if engine == "openpyxl":
        yield from _iter_openpyxl_rows(path, sheet_name)
        return
    if engine == "xlrd":
        yield from _iter_xlrd_rows(path, sheet_name)
        return
    raise ValueError(f"Unknown engine '{engine}'.")


def read_excel_bulk(path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
    path, engine = _prepare_read(path)
    df = pd.read_excel(path, sheet_name=sheet_name or 0, engine=engine)
    return _df_replace_nans_inplace(df)


def read_excel_bigdata(
    path: Path,
    sheet_name: Optional[str] = None,
    npartitions: int = 8,
) -> dd.DataFrame:
    path, engine = _prepare_read(path)
    pdf = pd.read_excel(path, sheet_name=sheet_name or 0, engine=engine)
    pdf = _df_replace_nans_inplace(pdf)
    npartitions = max(1, min(npartitions, max(len(pdf), 1)))
    return dd.from_pandas(pdf, npartitions=npartitions)


def _normalize_to_dataframe(
    data: Union[pd.DataFrame, Sequence[Dict[str, object]]],
) -> pd.DataFrame:
    return (
        data
        if isinstance(data, pd.DataFrame)
        else pd.DataFrame(list(data) if data else [])
    )


def _first_row_empty(ws) -> bool:
    """Check if the first row is empty or missing (header not yet written)."""
    if ws.max_row == 0:
        return True
    first = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not first:
        return True
    row_vals = first[0]
    return not any(v not in (None, "") for v in row_vals)


def _open_or_create_wb_ws(path: Path, sheet_name: Optional[str]):
    """
    Open workbook and target worksheet; create them if missing.
    Ensure target worksheet is the first sheet (so pandas sees header correctly).
    """
    from openpyxl import load_workbook, Workbook

    target = sheet_name or SHEET_DEFAULT
    if path.exists():
        wb = load_workbook(path)
        if target in wb.sheetnames:
            ws = wb[target]
        else:
            ws = wb.create_sheet(title=target)
        try:
            wb.move_sheet(ws, offset=-len(wb.worksheets))
        except Exception:
            wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws)))
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = target
    return wb, ws


def write_excel_row(
    path: Path, row: Dict[str, object], sheet_name: Optional[str] = None
) -> None:
    """
    Strict append of a single row using openpyxl only. No fallbacks.
    Ensures header is in the very first row (so pandas sees it).
    """
    path, engine = _prepare_write(path)
    if engine != "openpyxl":
        raise ValueError(
            "Appending rows is only supported for .xlsx/.xlsm via openpyxl."
        )

    wb, ws = _open_or_create_wb_ws(path, sheet_name)

    if _first_row_empty(ws):
        keys = list(row.keys())
        for idx, key in enumerate(keys, start=1):
            ws.cell(row=1, column=idx, value=key)
        for idx, key in enumerate(keys, start=1):
            ws.cell(row=2, column=idx, value=row.get(key))
    else:
        header = [cell.value for cell in ws[1]]
        ws.append([row.get(k) for k in header])

    wb.save(path)


def write_excel_bulk(
    path: Path,
    data: Union[pd.DataFrame, List[Dict[str, object]]],
    sheet_name: Optional[str] = None,
) -> None:
    path, engine = _prepare_write(path)
    df = _normalize_to_dataframe(data)
    df.to_excel(
        path, sheet_name=sheet_name or SHEET_DEFAULT, index=False, engine=engine
    )


def write_excel_bigdata(
    path: Path,
    data: dd.DataFrame,
    sheet_name: Optional[str] = None,
) -> None:
    path, engine = _prepare_write(path)
    pdf = data.compute()
    pdf.to_excel(
        path, sheet_name=sheet_name or SHEET_DEFAULT, index=False, engine=engine
    )
